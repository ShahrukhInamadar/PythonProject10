# import pytest
# import openpyxl
# from selenium.common.exceptions import NoSuchElementException
# from pages.LoginPage import LoginPage
# from tests.BaseTest import BaseTest
#
# wb = openpyxl.load_workbook("testdata.xlsx")
# sheet = wb.active
#
# def get_login_data():
#     headings = {}
#     data = []
#     for col in range(1, sheet.max_column + 1):
#         heading_name = sheet.cell(row=1, column=col).value
#         if heading_name:
#             headings[heading_name.strip().lower()] = col
#     for row in range(2, sheet.max_row + 1):
#         username = sheet.cell(row=row, column=headings["username"]).value
#         password = sheet.cell(row=row, column=headings["password"]).value
#         result_col = headings["result"]
#         data.append((username, password, row, result_col))
#     return data
#
# class TestLogin(BaseTest):
#
#     @pytest.mark.parametrize("username,password,row,result_col", get_login_data())
#     def test_login(self, init_driver, username, password, row, result_col):
#         loginpage = LoginPage(init_driver)
#         loginpage.enter_user_name(username)
#         loginpage.enter_the_password(password)
#         loginpage.click_on_submit()
#
#         print(f"Writing result at row {row}, column {result_col}")  # ✅ DEBUG PRINT
#
#         try:
#             if init_driver.current_url == "https://practicetestautomation.com/logged-in-successfully/":
#                 sheet.cell(row=row, column=result_col).value = "Pass"
#                 print(f"✅ {username} login success")
#             else:
#                 sheet.cell(row=row, column=result_col).value = "Fail"
#                 print(f"❌ {username} login failed")
#         except NoSuchElementException:
#             sheet.cell(row=row, column=result_col).value = "Fail"
#             print(f"❌ {username} login failed - NoSuchElement")
#
#         wb.save("testdata.xlsx")  # ✅ Ensure this line executes


import pytest
import openpyxl
from selenium.common.exceptions import NoSuchElementException
from pages.LoginPage import LoginPage
from tests.BaseTest import BaseTest


def get_login_data():
    wb = openpyxl.load_workbook("testdata.xlsx")
    sheet = wb.active
    headings = {}
    data = []

    for col in range(1, sheet.max_column + 1):
        heading_name = sheet.cell(row=1, column=col).value
        if heading_name:
            headings[heading_name.strip().lower()] = col

    for row in range(2, sheet.max_row + 1):
        username = sheet.cell(row=row, column=headings["username"]).value
        password = sheet.cell(row=row, column=headings["password"]).value
        result_col = headings["result"]
        data.append((username, password, row, result_col))

    wb.close()
    return data


class TestLogin(BaseTest):

    @pytest.mark.parametrize("username,password,row,result_col", get_login_data())
    def test_login(self, init_driver, username, password, row, result_col):
        loginpage = LoginPage(init_driver)
        loginpage.enter_user_name(username)
        loginpage.enter_the_password(password)
        loginpage.click_on_submit()

        wb = openpyxl.load_workbook("testdata.xlsx")
        sheet = wb.active
        expected_result = sheet.cell(row=row, column=result_col).value

        try:
            current_url = init_driver.current_url.strip("/")
            success_url = "https://practicetestautomation.com/logged-in-successfully".strip("/")

            if current_url == success_url:
                actual_result = "Pass"
                print(f"✅ {username} login success")
            else:
                actual_result = "Fail"
                print(f"❌ {username} login failed - wrong URL")

        except NoSuchElementException as e:
            actual_result = "Fail"
            print(f"❌ {username} login failed - NoSuchElement: {e}")

        # Save actual result in the next column
        sheet.cell(row=row, column=result_col + 1).value = actual_result
        wb.save("testdata.xlsx")
        wb.close()

        # Final assertion
        assert actual_result == expected_result, f"Expected: {expected_result}, Got: {actual_result}"
